import openpyxl
from easydbo import constant
from easydbo.output.log import Log
from easydbo.excel.newdata import NewData


class ExcelOperation:
    def __init__(self, config, excel_path, tblexl):
        self.config = config
        self.excel_path = excel_path
        self.sheet = tblexl.name
        self.columns = tblexl.columns
        self.primary_idx = tblexl.pkidx
        self.date_column = tblexl.get_cols_date()
        #
        self.data = self._load()  # self.data[row][column]

    def _load(self):
        wb = openpyxl.load_workbook(self.excel_path)
        sheetnum = wb.sheetnames.index(self.sheet)
        ws = wb.worksheets[sheetnum]
        # Header
        idx_valid = []
        idx_date = []
        for row in ws.iter_rows(min_row=1, max_row=1):
            for idx, cell in enumerate(row):
                if cell.value in self.columns:
                    idx_valid.append(idx)
                if cell.value in self.date_column:
                    idx_date.append(idx)
        if len(set(idx_valid)) != len(idx_valid):
            Log.error(f'Duplicate column names exist in {self.sheet}(sheet)')
        # Data
        return NewData(idx_valid, idx_date).normalize(ws, sheet=self.sheet)

    def check_unique(self, idxes):
        max_len = len(self.data)
        for idx in idxes:
            d = [d[idx] for d in self.data]
            if len(set(d)) != max_len:
                Log.error(f'{self.columns[idx]}(column) of {self.sheet}(sheet) must have unique elements')

    def check_null(self, idxes):
        for idx in idxes:
            for d in self.data:
                if d == constant.NAN_STR:
                    Log.error(f'{self.columns[idx]}(column) of {self.sheet}(sheet) must not be empty')

    def col_to_idx(self, cols):
        return [self.columns.index(c) for c in cols]

    def get_data(self):
        return self.data

    def select(self, cols):
        col_idx = self.col_to_idx(cols)
        data = [[d[i] for i in col_idx] for d in self.data]
        return data
