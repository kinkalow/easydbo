from easydbo.output.log import Log
from easydbo import constant
from datetime import datetime
import openpyxl

class NewData:
    def __init__(self, idx_valid, idx_date):
        self.idx_valid = idx_valid
        self.idx_date = idx_date
        #
        self.is_ws_changed = False

    def to_worksheet(self, data):
        '''
        Convert data to worksheet
        Parm data: 2D list
        '''
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['' for _ in range(len(data))])  # Dummy header
        for d in data:
            ws.append([d_ for i, d_ in enumerate(d) if i in self.idx_valid])
        for i in self.idx_date:
            for w in ws.iter_rows(min_row=2):
                try:
                    w[i].value = datetime.strptime(w[i].value, constant.DATE_FORMAT)
                except ValueError:
                    Log.error(f'"{w[i].value}" must be changed to match the following format "{constant.DATE_FORMAT}"')
        self.is_ws_changed = True
        return ws

    def normalize(self, ws, sheet=None):
        data = []
        for row in ws.iter_rows(min_row=2):
            datum = []
            is_empty_row = True
            for idx in self.idx_valid:
                cell = row[idx]
                value = cell.value
                if value is None:
                    value = constant.NAN_STR
                else:
                    is_empty_row = False
                    if idx in self.idx_date:
                        if not cell.is_date:
                            cell_sheet = f' cell={cell.coordinate} sheet={self.sheet}' if sheet else ''
                            Log.error(f'Not date type: value={cell.value}{cell_sheet}')
                        value = str(value.strftime(constant.DATE_FORMAT))
                    elif not isinstance(value, str):
                        value = str(value)
                datum.append(value)
            if not is_empty_row:
                data.append(datum)
        return data

class NewDataChecker:
    def __init__(self, dbop):
        self.dbop = dbop

    def unique(self, table, new_data):
        name = table.name
        idxes_uniq = table.get_idxes_uniq()
        cols_uniq = table.get_cols_uniq()
        if not idxes_uniq:
            return
        vals_uniq = self.dbop.select(name, cols_uniq)
        for idx, col in zip(idxes_uniq, cols_uniq):
            targets = [v[idx] for v in vals_uniq]
            for nd in new_data:
                if nd[idx] in targets:
                    Log.error(f'"{col}={nd[idx]}" must be unique')

    def null(self, table, new_data):
        idxes_null = table.get_idxes_null()
        for nd in new_data:
            for i, d in enumerate(nd):
                if i in idxes_null:
                    continue
                if d == constant.NAN_STR:
                    Log.error(f'"{table.columns[i]}" filed must not be null')
